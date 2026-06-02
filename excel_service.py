from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from scheduler_service import SHIFT_LABELS


def _safe_str(value):
    if value is None:
        return ""
    return str(value)


def _build_time_range(row):
    start_time = row["start_time"] if "start_time" in row.keys() else ""
    end_time = row["end_time"] if "end_time" in row.keys() else ""

    start_time = start_time or ""
    end_time = end_time or ""

    if start_time or end_time:
        return f"{start_time}-{end_time}".strip("-")
    return ""


def _auto_fit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)


def export_schedule_to_excel(rows, file_path):
    """
    將班表資料匯出為 Excel 檔案。

    rows:
        需支援以下欄位（sqlite3.Row 也可）：
        - id
        - work_date
        - secretary_name
        - shift_type
        - start_time
        - end_time
        - is_holiday_shift
        - source_type
        - remark
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "班表"

    headers = [
        "ID",
        "日期",
        "秘書",
        "班別",
        "時間",
        "假日班",
        "來源",
        "備註",
    ]

    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    for row in rows:
        shift_label = SHIFT_LABELS.get(row["shift_type"], row["shift_type"])
        time_range = _build_time_range(row)
        holiday_text = "是" if row["is_holiday_shift"] == 1 else "否"

        ws.append([
            row["id"],
            _safe_str(row["work_date"]),
            _safe_str(row["secretary_name"]),
            _safe_str(shift_label),
            time_range,
            holiday_text,
            _safe_str(row["source_type"]) if "source_type" in row.keys() else "",
            _safe_str(row["remark"]) if "remark" in row.keys() else "",
        ])

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.border = thin_border
            if cell.column in (1, 2, 4, 5, 6, 7):
                cell.alignment = center_alignment
            else:
                cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    _auto_fit_columns(ws)
    wb.save(file_path)